import os.path
import openpyxl

def extract_data(file, hoid_output, static_output):
    # target file open
    # 하나 배웠다, data_only안하면 cell에 써져있는 그대로 가져오고
    # data only하면 출력되는 모습 그대로 가져오는 듯 하다
    xlsx = openpyxl.load_workbook(filename=file, data_only=True)
    sheet = xlsx['Sheet']
    #
    #static_first = ["file_name", "ALL", "평민 이상", "노비", "공백", "x포함"]
    static_count = [file, 0, 0, 0, 0, 0]

    for i in sheet.rows:
        # 성, 명이 있는 주호 중 주성명이 비어있는 사람만 추출
        if i[22].value != None and "x" not in i[22].value and i[23].value != None and "x" not in i[23].value and i[42].value == None:
            # hoid
            a = [str(i[4].value).split('.')[0] + str(i[2].value).split('.')[0] + "-" + str(i[8].value.split('.')[0]) \
                 + str(i[11].value).split('.')[0] + str(i[13].value).split('.')[0]]
            # 직역은 변경될 가능성이 있으니 지금은 배제하도록 하자
            #a.append(i[19].value) # 직역(한글)
            a.append(i[22].value) # 성
            a.append(i[23].value) # 명
            if type(i[24].value) == int: # 출생년도
                a.append(i[4].value - i[24].value)
            else:
                a.append(i[24].value)
            a.append(i[26].value)   # 간지(한글)
            a.append(i[42].value)   # 주성명, 노비인 경우에 사용
            a.append(i[45].value)   # 부명(한자)
            a.append(i[46].value)   # 부명(한글)
            a.append(i[49].value)   # 모명(한자)
            a.append(i[50].value)   # 모명(한글)
            a.append(i[54].value)   # 조명(한자)
            a.append(i[55].value)   # 조명(한글)
            a.append(i[58].value)   # 증조명(한자)
            a.append(i[59].value)   # 증조명(한글)
            a.append(i[62].value)   # 외조명(한글)
            a.append(i[63].value)   # 외조명(한글)
            hoid_output.append(a)

        # static
        static_count[1] += 1
        job = i[18].value
        if job == None:
            static_count[4] += 1
        elif '奴' in job or '婢' in job:
            static_count[3] += 1
        elif 'x' in job:
            static_count[5] += 1
        else:
            static_count[2] += 1
    static_output.append(static_count)

def main():
    # 작업 위치 및 저장 위치 정의
    work_dir = os.getcwd()
    print(work_dir + "현재 작업 위치")
    id_dir = work_dir + '\\' + "output_식별데이터" + '\\'

    # save file open
    xlsx1 = openpyxl.Workbook()
    hoid_output = xlsx1.active

    # statics file open
    xlsx2 = openpyxl.Workbook()
    static_output = xlsx2.active
    static_first = ["file_name", "ALL", "평민 이상", "노비", "공백", "x포함"]
    static_output.append(static_first)

    # output 저장 폴더 확인
    if os.path.isdir(id_dir) == 0:
        os.mkdir(id_dir)

    files = os.listdir(work_dir)

    for file in files:
        if len(file.split('.')) == 2:
            if file.split('.')[1] == 'xlsx':
                print(file)
                extract_data(file, hoid_output, static_output)

    save_file_mho = id_dir + "mho_id.xlsx"
    save_file_static = id_dir + "statics_mho.xlsx"
    xlsx1.save(save_file_mho)
    xlsx2.save(save_file_static)

main()