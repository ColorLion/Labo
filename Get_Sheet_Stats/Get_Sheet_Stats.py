import os
import pyodbc
import openpyxl

def conn_excel(ex_file):
    #excel_conn
    ex_conn_str = (
        'Driver={Microsoft Excel Driver (*.xls, *.xlsx, *.xlsm, *.xlsb)};'+
        'DBQ='+ex_file
        )
    ex_cnxn = pyodbc.connect(ex_conn_str, autocommit=True)
    ex_cur = ex_cnxn.cursor()

    return ex_cur

def conn_db(db_file, result):
    # db_conn
    dbconn_string = (
        'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'+
        'DBQ='+db_file
        )
    db_conn = pyodbc.connect(dbconn_string)
    db_cur = db_conn.cursor()

    # insert excel data
    db_query = 'insert into 통합 (region_id, 성, 명, 출생년도, 간지) values (?, ?, ?, ?, ?)'
    db_cur.executemany(db_query, result)

    return db_cur

    '''
    db_cur.rollback()
    #db_cur.commit()
    db_cur.close()
    '''

def extract_excel_data(ex_cur):
    #get sheet name
    for row in ex_cur.tables():
        sheet = row.table_name

    #get sheet data
    ex_query = 'select * from ['+sheet+']'
    ex_results = ex_cur.execute(ex_query).fetchall()

    return ex_results

def extract_access_data(ex, db_cur, mov):
    # report 파일 생성
    ex_report = ex.split('.')[0] + '_stats_note.xlsx'
    report2 = openpyxl.Workbook()

    # make sheet frame
    frame1 = ['성', '명', '출생년도', '간지', 'region_id']
    frame2 = ['region_id', '성', '명', '출생년도', '간지']
 
    # 추출
    sql1 = 'select * from 1_중복_2'
    ws1 = report2.create_sheet("1.중복_2")
    ws1.append(frame1)
    ws1.column_dimensions['E'].width = 12
    t = db_cur.execute(sql1)
    for i in t:
        item = list(i)
        ws1.append(item)

    sql2 = 'select * from 1_중복_3'
    ws2 = report2.create_sheet("1.중복_3")
    ws2.append(frame1)
    ws2.column_dimensions['E'].width = 12
    t = db_cur.execute(sql2)
    for i in t:
        item = list(i)
        ws2.append(item)

    sql3 = 'select * from 1_중복_4'
    ws3 = report2.create_sheet("1.중복_4")
    ws3.append(frame1)
    ws3.column_dimensions['E'].width = 12
    t = db_cur.execute(sql3)
    for i in t:
        item = list(i)
        ws3.append(item)

    sql4 = 'select * from 1_중복_5'
    ws4 = report2.create_sheet("1.중복_5")
    ws4.append(frame1)
    ws4.column_dimensions['E'].width = 12
    t = db_cur.execute(sql4)
    for i in t:
        item = list(i)
        ws4.append(item)

    sql5 = 'select * from 1_중복_over'
    ws5 = report2.create_sheet("1.중복_over")
    ws5.append(frame2)
    ws5.column_dimensions['A'].width = 12
    t = db_cur.execute(sql5)
    for i in t:
        item = list(i)
        ws5.append(item)

    sql6 = 'select * from 2_명_공백'
    ws6 = report2.create_sheet("2.명_공백")
    ws6.append(frame2)
    ws6.column_dimensions['A'].width = 12
    t = db_cur.execute(sql6)
    for i in t:
        item = list(i)
        ws6.append(item)

    sql7 = 'select * from 3_명_x'
    ws7 = report2.create_sheet("3.명_x")
    ws7.append(frame2)
    ws7.column_dimensions['A'].width = 12
    t = db_cur.execute(sql7)
    for i in t:
        item = list(i)
        ws7.append(item)

    sql8 = 'select * from 4_출생년도_공백'
    ws8 = report2.create_sheet("4.출생년도_공백")
    ws8.append(frame2)
    ws8.column_dimensions['A'].width = 12
    t = db_cur.execute(sql8)
    for i in t:
        item = list(i)
        ws8.append(item)

    sql9 = 'select * from 5_간지_공백'
    ws9 = report2.create_sheet("5.간지_공백")
    ws9.append(frame2)
    ws9.column_dimensions['A'].width = 12
    t = db_cur.execute(sql9)
    for i in t:
        item = list(i)
        ws9.append(item)

    sql10 = 'select * from 6_간지_x'
    ws10 = report2.create_sheet("6.간지_x")
    ws10.append(frame2)
    ws10.column_dimensions['A'].width = 12
    t = db_cur.execute(sql10)
    for i in t:
        item = list(i)
        ws10.append(item)

    # 처음 생성되는 sheet 삭제
    remove = report2.get_sheet_by_name('Sheet')
    report2.remove_sheet(remove)

    # save report file
    save = mov + '\\' + ex_report
    report2.save(save)

def get_db_stats(db_cur, ex, rx1):
    year = extract_year(ex)
    stats = [year]
    overlap = 0

    sql1 = 'select count(*) from 통합'
    t = db_cur.execute(sql1)
    for i in t:
        stats.append(i[0])

    sql2 = 'select count(*) from 1_중복_2'
    t = db_cur.execute(sql2)
    for i in t:
        stats.append(i[0])
        overlap = overlap + i[0]

    sql3 = 'select count(*) from 1_중복_3'
    t = db_cur.execute(sql3)
    for i in t:
        stats.append(i[0])
        overlap = overlap + i[0]

    sql4 = 'select count(*) from 1_중복_4'
    t = db_cur.execute(sql4)
    for i in t:
        stats.append(i[0])
        overlap = overlap + i[0]

    sql5 = 'select count(*) from 1_중복_5'
    t = db_cur.execute(sql5)
    for i in t:
        stats.append(i[0])
        overlap = overlap + i[0]

    sql6 = 'select count(*) from 1_중복_over'
    t = db_cur.execute(sql6)
    for i in t:
        stats.append(i[0])
        overlap = overlap + i[0]
        stats.append(overlap)

    sql7 = 'select count(*) from 2_명_공백'
    t = db_cur.execute(sql7)
    for i in t:
        stats.append(i[0])

    sql8 = 'select count(*) from 3_명_x'
    t = db_cur.execute(sql8)
    for i in t:
        stats.append(i[0])

    sql9 = 'select count(*) from 4_출생년도_공백'
    t = db_cur.execute(sql9)
    for i in t:
        stats.append(i[0])

    sql10 = 'select count(*) from 5_간지_공백'
    t = db_cur.execute(sql10)
    for i in t:
        stats.append(i[0])

    sql11 = 'select count(*) from 6_간지_x'
    t = db_cur.execute(sql11)
    for i in t:
        stats.append(i[0])

    rx1.append(stats)
    rx1.column_dimensions['A'].width = 13

def extract_year(ex):
    year = ex.split('_')[0]
    return year

def main():
    now_dir = os.getcwd()
    folder = os.getcwd()
    dir_name = 'output_stats_note'

    if os.path.isdir(dir_name) == False:
        os.mkdir(dir_name)
        mov = folder + '\\' + dir_name + '\\'
    else:
        mov = folder + '\\' + dir_name + '\\'

    filenames = os.listdir(folder)

     # find db file
    db_file = now_dir + '\\' + 'stats.accdb'

    #make report file
    report1 = openpyxl.Workbook()
    rx1 = report1.active
    total_frame = ['파일', '레코드 수', '중복_2', '중복_3', '중복_4', '중복_5', '중복_over', 'sum_중복', '명_공백', '명_x', '출생년도_공백', '간지_공백', '간지_x']
    rx1.append(total_frame)

    # find excel file and work start
    filenames = os.listdir(now_dir)
    for filename in filenames:
        if len(filename.split('.')) == 2:
            # except file
            if filename.split('_')[0] == 'report':
                print("[-] With out : " + filename)
            # target file
            elif filename.split('.')[1] == 'xlsx':
                # filename
                ex = filename
                ex_file = now_dir + '\\' + filename
                print(ex_file)

                # excel
                ex_cur = conn_excel(ex_file)
                result = extract_excel_data(ex_cur)

                # db
                db_cur = conn_db(db_file, result)
                get_db_stats(db_cur, ex, rx1)
                extract_access_data(ex, db_cur, mov)

                # close
                db_cur.rollback()
                db_cur.close()

    report1_name = '개별시트통계정보.xlsx'
    report1.save(report1_name)

main()