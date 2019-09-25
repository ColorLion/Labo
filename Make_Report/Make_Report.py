import os.path
import xlrd
import openpyxl


def make_report(file, first_row, report_rows, mainho_rows, blank_rows):
    # 단성호적 파일 open
    target = xlrd.open_workbook(file, formatting_info=True)
    ws = target.sheet_by_index(0)

    # 배열 내 연도의 위치 확인
    year = file.split('-')[0]
    j = first_row.index(year)

    # 위치 저장
    location = file.split('.')[0][7:]
    location_code = file.split('.')[0][5:7]
    # print(location)
    # print(location_code)

    if location_code == '01':
        # 첫번 째 행에 위치 저장
        report_rows[0][0] = location
        mainho_rows[0][0] = location
        blank_rows[0][0] = location

        # 연도별 인물의 수 저장
        report_rows[0][j] = ws.nrows - 1

        # 연도별 주호의 수 저장
        mainho_rows[0][j] = main_ho_counting(ws)

        # 연도별 가족 구성원의 성, 본관의 빈칸 데이터
        blank_rows[0][j] = blank_counting(ws)

    elif location_code == '02':
        # 첫번 째 행에 위치 저장
        report_rows[1][0] = location
        mainho_rows[1][0] = location
        blank_rows[1][0] = location

        # 연도별 인물의 수 저장
        report_rows[1][j] = ws.nrows - 1

        # 연도별 주호의 수 저장
        mainho_rows[1][j] = main_ho_counting(ws)

        # 연도별 가족 구성원의 성, 본관의 빈칸 데이터
        blank_rows[1][j] = blank_counting(ws)

    elif location_code == '03':
        # 첫번 째 행에 위치 저장
        report_rows[2][0] = location
        mainho_rows[2][0] = location
        blank_rows[2][0] = location

        # 연도별 인물의 수 저장
        report_rows[2][j] = ws.nrows - 1

        # 연도별 주호의 수 저장
        mainho_rows[2][j] = main_ho_counting(ws)

        # 연도별 가족 구성원의 성, 본관의 빈칸 데이터
        blank_rows[2][j] = blank_counting(ws)

    elif location_code == '04':
        # 첫번 째 행에 위치 저장
        report_rows[3][0] = location
        mainho_rows[3][0] = location
        blank_rows[3][0] = location

        # 연도별 인물의 수 저장
        report_rows[3][j] = ws.nrows - 1

        # 연도별 주호의 수 저장
        mainho_rows[3][j] = main_ho_counting(ws)

        # 연도별 가족 구성원의 성, 본관의 빈칸 데이터
        blank_rows[3][j] = blank_counting(ws)

    elif location_code == '05':
        # 첫번 째 행에 위치 저장
        report_rows[4][0] = location
        mainho_rows[4][0] = location
        blank_rows[4][0] = location

        # 연도별 인물의 수 저장
        report_rows[4][j] = ws.nrows - 1

        # 연도별 주호의 수 저장
        mainho_rows[4][j] = main_ho_counting(ws)

        # 연도별 가족 구성원의 성, 본관의 빈칸 데이터
        blank_rows[4][j] = blank_counting(ws)

    elif location_code == '06':
        # 첫번 째 행에 위치 저장
        report_rows[5][0] = location
        mainho_rows[5][0] = location
        blank_rows[5][0] = location

        # 연도별 인물의 수 저장
        report_rows[5][j] = ws.nrows - 1

        # 연도별 주호의 수 저장
        mainho_rows[5][j] = main_ho_counting(ws)

        # 연도별 가족 구성원의 성, 본관의 빈칸 데이터
        blank_rows[5][j] = blank_counting(ws)

    elif location_code == '07':
        # 첫번 째 행에 위치 저장
        report_rows[6][0] = location
        mainho_rows[6][0] = location
        blank_rows[6][0] = location

        # 연도별 인물의 수 저장
        report_rows[6][j] = ws.nrows - 1

        # 연도별 주호의 수 저장
        mainho_rows[6][j] = main_ho_counting(ws)

        # 연도별 가족 구성원의 성, 본관의 빈칸 데이터
        blank_rows[6][j] = blank_counting(ws)

    elif location_code == '08':
        # 첫번 째 행에 위치 저장
        report_rows[7][0] = location
        mainho_rows[7][0] = location
        blank_rows[7][0] = location

        # 연도별 인물의 수 저장
        report_rows[7][j] = ws.nrows - 1

        # 연도별 주호의 수 저장
        mainho_rows[7][j] = main_ho_counting(ws)

        # 연도별 가족 구성원의 성, 본관의 빈칸 데이터
        blank_rows[7][j] = blank_counting(ws)


def main_ho_counting(ws):
    result = 0

    for i in range(ws.nrows):
        row = ws.row_values(i)
        if "주호" in row[17]:
            result += 1

    return result


def blank_counting(ws):
    result = 0

    for i in range(ws.nrows):
        row = ws.row_values(i)
        if "子" in row[16] or "女" in row[16] or "父" in row[16]:
            result += 1

    return result


def make_first_row(first_row, work_dir):
    files = os.listdir(work_dir)

    for file in files:
        if len(file.split('.')) == 2:
            if file.split('.')[1] == 'xls':
                if file.split('-')[0] not in first_row:
                    first_row.append(file.split('-')[0])

    return first_row


def main():
    # 작업 위치 및 저장 위치 정의
    work_dir = os.getcwd()
    print(work_dir + "현재 작업 위치")
    report_dir = work_dir + '\\' + "output_report" + '\\'

    # output 저장 폴더 확인
    if os.path.isdir(report_dir) == 0:
        os.mkdir(report_dir)

    files = os.listdir(work_dir)

    # 첫 행을 위한 리스트
    first_row = ['지역/연도']

    # 첫번째 행 만들기
    first_row = make_first_row(first_row, work_dir)
    print(first_row)

    # 내용물을 채우기 위한 리스트
    report_rows = [[0] * len(first_row) for i in range(8)]  # 사람 수
    mainho_rows = [[0] * len(first_row) for i in range(8)]  # 주호 수
    blank_rows = [[0] * len(first_row) for i in range(8)]  # 구성원의 빈칸 수

    for file in files:
        if len(file.split('.')) == 2:
            if file.split('.')[1] == 'xls':
                print(file)
                make_report(file, first_row, report_rows, mainho_rows, blank_rows)

    # report 저장용 xlsx 파일
    xlsx = openpyxl.Workbook()

    # 연도별 파일의 rows 수 저장
    report_1 = xlsx.active
    report_1.title = 'Rows'
    report_1.append(first_row)
    for i in range(8):
        print(report_1.append(report_rows[i]))

    # 연도별 파일의 주호 수 저장
    report_2 = xlsx.create_sheet()
    report_2.title = 'main_ho'
    report_2.append(first_row)
    for i in range(8):
        print(report_2.append(mainho_rows[i]))

    # 연도별 파일의 가족 구성원이기 때문에 성, 본관 등이 비어 있는 데이터의 수
    report_3 = xlsx.create_sheet()
    report_3.title = 'blank_column'
    report_3.append(first_row)
    for i in range(8):
        print(report_3.append(blank_rows[i]))

    # report 파일 저장
    xlsx.save(report_dir + '\\' + "report.xlsx")

main()