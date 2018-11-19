import os
import pyodbc
import openpyxl

def conn_excel(ex_file):
    #excel_conn a
    ex_conn_str = (
        'Driver={Microsoft Excel Driver (*.xls, *.xlsx, *.xlsm, *.xlsb)};'+
        'DBQ='+ex_file
        )
    ex_cnxn = pyodbc.connect(ex_conn_str, autocommit=True)
    ex_cur = ex_cnxn.cursor()

    return ex_cur

def conn_db(db_file, result):
    # db_conn
    dbconn_string = (
        'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'+
        'DBQ='+db_file
        )
    db_conn = pyodbc.connect(dbconn_string)
    db_cur = db_conn.cursor()

    # insert excel data
    db_query = 'insert into 통합 (region_id, 성, 명, 출생년도, 간지) values (?, ?, ?, ?, ?)'
    db_cur.executemany(db_query, result)

    return db_cur

def extract_excel_data(ex_cur):
    #get sheet name
    for row in ex_cur.tables():
        sheet = row.table_name

    #get sheet data
    ex_query = 'select * from ['+sheet+']'
    ex_results = ex_cur.execute(ex_query).fetchall()

    return ex_results

def get_db_stats(db_cur, ex, rx1):
    year = extract_year(ex)
    stats = [year]
    # 동일인물 sum
    overlap_1 = 0
    # 식별인물 + 단일인물
    overlap_2 = 0

    sql1 = 'select count(*) from 통합'
    t = db_cur.execute(sql1)
    for i in t:
        stats.append(i[0])

    sql2 = 'select count(*) from 동일_2'
    t = db_cur.execute(sql2)
    for i in t:
        stats.append(i[0])
        overlap_1 = overlap_1 + i[0]

    sql3 = 'select count(*) from 동일_3'
    t = db_cur.execute(sql3)
    for i in t:
        stats.append(i[0])
        overlap_1 = overlap_1 + i[0]

    sql4 = 'select count(*) from 동일_4'
    t = db_cur.execute(sql4)
    for i in t:
        stats.append(i[0])
        overlap_1 = overlap_1 + i[0]

    sql5 = 'select count(*) from 동일_5'
    t = db_cur.execute(sql5)
    for i in t:
        stats.append(i[0])
        overlap_1 = overlap_1 + i[0]

    sql6 = 'select count(*) from 동일_over'
    t = db_cur.execute(sql6)
    for i in t:
        stats.append(i[0])
        overlap_1 = overlap_1 + i[0]
        stats.append(overlap_1)

    sql7 = 'select count(*) from 식별인물'
    t = db_cur.execute(sql7)
    for i in t:
        stats.append(i[0])
        overlap_2 = overlap_2 + i[0]

    sql8 = 'select count(*) from 단일인물'
    t = db_cur.execute(sql8)
    for i in t:
        stats.append(i[0])
        overlap_2 = overlap_2 + i[0]
        stats.append(overlap_2)

    rx1.append(stats)
    #rx1.column_dimensions['A'].width = 13

def get_bind_info(db_cur):
    report3 = openpyxl.Workbook()
    report4 = openpyxl.Workbook()

    # sheet
    wss1 = report3.create_sheet("통합")
    wss2 = report3.create_sheet("ID")
    wss3 = report3.create_sheet("Search")
    rx4 = report4.active

    # columns
    frame1 = ['region_id', '성', '명', '출생년도', '간지']
    wss1.append(frame1)
    rx4.append(frame1)
    frame2 = ['성', '명', '출생년도', '간지', 'global_id']
    wss2.append(frame2)
    frame3 = ['global_id', 'region_id', '성', '명', '출생년도', '간지']
    wss3.append(frame3)

    sql1 = 'select * from 통합'
    wss1.column_dimensions['A'].width = 12
    rx4.column_dimensions['A'].width = 12
    t = db_cur.execute(sql1)
    for i in t:
        item = list(i)
        wss1.append(item)
        rx4.append(item)

    sql2 = 'select * from ID'
    wss2.column_dimensions['E'].width = 12
    t = db_cur.execute(sql2)
    for i in t:
        item = list(i)
        wss2.append(item)

    sql3 = 'select * from Search'
    wss3.column_dimensions['A'].width = 12
    wss3.column_dimensions['B'].width = 12
    t = db_cur.execute(sql3)
    for i in t:
        item = list(i)
        wss3.append(item)

    # 처음 생성되는 sheet 삭제
    remove = report3.get_sheet_by_name('Sheet')
    report3.remove_sheet(remove)

    report3.save("bind_info.xlsx")
    report4.save("combine.xlsx")

def make_globalID(db_cur):
    sql1 = 'insert into ID select * from make_ID_1'
    db_cur.execute(sql1)

    sql2 = 'insert into ID select * from make_ID_2'
    db_cur.execute(sql2)

    sql3 = 'insert into Search select * from make_Search'
    db_cur.execute(sql3)

    db_cur.commit()

def extract_year(ex):
    year = ex.split('_')[0]
    return year

def main():
    now_dir = os.getcwd()

    #make report file
    report1 = openpyxl.Workbook()
    rx1 = report1.active
    total_frame = ['통합', '누적 인물 레코드 수', '동일_2', '동일_3', '동일_4', '동일_5', '동일_over', '동일 인물', '식별 인물', '단일 인물', '식별 + 단일']
    rx1.append(total_frame)
    
    # find db file
    db_file = now_dir + '\\' + 'combine.accdb'

    # find excel file and work start
    filenames = os.listdir(now_dir)
    for filename in filenames:
        if len(filename.split('.')) == 2:
            # except file
            if filename.split('_')[0] == "report":
                print("[-] With out : " + filename)
            # target file
            elif filename.split('.')[1] == 'xlsx':
                # filename
                ex = filename
                ex_file = now_dir + '\\' + filename
                print(ex_file)

                # excel
                ex_cur = conn_excel(ex_file)
                result = extract_excel_data(ex_cur)

                # db
                db_cur = conn_db(db_file, result)
                get_db_stats(db_cur, ex, rx1)

                db_cur.commit()

    # save report
    report1_name = '통합과정통계정보.xlsx'
    rx1.column_dimensions['K'].width = 12
    report1.save(report1_name)

    # make global_id, search table
    make_globalID(db_cur)
    get_bind_info(db_cur)
    
    db_cur.close()

main()