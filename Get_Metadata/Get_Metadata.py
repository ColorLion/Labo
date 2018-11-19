import xlrd
import os
import openpyxl

# global
# meta - column 순번, data type, 길이, null의 수
meta = [[]]
length = []
null = 0

# column을 추출할 함수
def xls_column(frow):
    column = [[0, frow[0]]]
    for i in range(1, len(frow)):
        column.append([i, frow[i]])
    return column

# column의 데이터 타입, 타입 별 데이터 수를 추출할 함수

def xls_metadata(orow, row):
    if row == 1:
        meta.clear()
        for i in range(len(orow)):
            # 시트별 내용
            if len(str(orow[i])) == 0:
                meta.append([i, type(orow[i]), len(str(orow[i])), 1])
            else:
                meta.append([i, type(orow[i]), len(str(orow[i])), 0])
    else:
        for i in range(len(orow)):
            # 시트별 내용
            if meta[i][2] < len(str(orow[i])):
                meta[i][2] = len(str(orow[i]))
            if len(str(orow[i])) == 0:
                meta[i][3] += 1

    return meta

def open_xls(excel, filename, xlsx):
    # xls 파일을 열기 위해 사용
    sheets = excel.nsheets
    fn = filename.split('.')[0]

    for sheet in range(sheets):
        sheet_name = excel.sheet_by_index(sheet)
        sn = excel.sheet_names()

        # 시트 내 데이터가 있다면 데이터 집계
        if sheet_name.nrows != 0:
            # first row
            first_col = xls_column(sheet_name.row_values(0))

            # other rows
            for row in range(1, sheet_name.nrows):
                xls_metadata(sheet_name.row_values(row), row)

            t = []
            # length
            for i in range(sheet_name.ncols):
                t.append(meta[i][2])
            length.append(t)

            # save
            sh = fn + '_' + sn[sheet]
            xlsx = save_as_sheet(xlsx, first_col, meta, sh)

    return xlsx

def open_xlsx(excel, filename, xlsx):
    print("open_xlsx")

def save_as_sheet(xlsx, first_col, meta, sheet_name):
    # xlsx 오픈
    #xlsx = openpyxl.Workbook()
    # 시트 만듦
    ax = xlsx.create_sheet(title=sheet_name)

    # column 넓이
    ax.column_dimensions['A'].width = 12
    ax.column_dimensions['B'].width = 14
    ax.column_dimensions['C'].width = 11
    ax.column_dimensions['D'].width = 11

    # xlsx에 저장
    first = ['항목명', 'data_type', 'max_length', 'null_count']
    ax.append(first)
    for i in range(len(first_col)):
        data = [first_col[i][1], str(meta[i][1]), meta[i][2], meta[i][3]]
        ax.append(data)

    meta.clear()
    return xlsx

def main():
    # 작업한 파일 수
    cnt = 0
    xl_file = ['항목명', 'data_type', 'max_length']

    # 작업 위치 지정
    folder = os.getcwd()
    notice = "현재 작업 위치 : "
    print(notice + folder)

    # 출력물 위치 지정
    folder = os.getcwd()
    dir_name = 'output_metadata'
    if os.path.isdir(dir_name) == False:
        os.mkdir(dir_name)
    mov = folder + '\\' + dir_name + '\\'

    insert = input("file name: ")
    name = mov + insert + '.xlsx'

    # xlsx 오픈
    xlsx = openpyxl.Workbook()

    # length 저장을 해보도록하자
    sh = xlsx.create_sheet(title='report')

    # 폴더 탐색
    filenames = os.listdir(folder)
    for filename in filenames:
        if len(filename.split('.')) == 2:
            if filename.split('.')[1] == 'xls':
                print(filename)
                xl_file.append(filename.split('.')[0])
                excel = xlrd.open_workbook(filename, formatting_info=True)
                xlsx = open_xls(excel, filename, xlsx)
                cnt = cnt + 1
            elif filename.split('.')[1] == 'xlsx':
                print(filename)
                xl_file.append(filename.split('.')[0])
                # excel = openpyxl

    output1 = []
    output2 = [[]]
    output2.clear()
    sh.append(xl_file)
    for j in range(len(length[0])):
        output1.clear()
        output1.append('')
        output1.append('')
        output1.append('')
        for i in range(cnt):
            output1.append(length[i][j])
        sh.append(output1)
    xlsx.save(filename=name)

main()