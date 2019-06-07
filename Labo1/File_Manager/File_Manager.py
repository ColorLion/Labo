import time
import xlrd
import openpyxl
import os
import sys

def search_xls(filename, fn, size):
	# open excel
	wb = xlrd.open_workbook(filename, formatting_info=True)

	# number of sheet
	number_sheet = wb.nsheets
	print('filename - ' + filename + ' / size - ' + str(size), file=fn)
	
	# print file infomation
	for i in range(number_sheet):
		# memory in sheet
		sheet_name = wb.sheet_by_index(i)
		sn = wb.sheet_names()
		print('	' + sn[i] + ' - ' + \
			'rows : ' + str(sheet_name.nrows) + \
			' cols : ' + str(sheet_name.ncols),file=fn)

def search_xlsx(filename, fn, size):
	# open excel
	wb = openpyxl.load_workbook(filename, read_only=True)

	# number of sheet
	ws = wb.get_sheet_names()
	number_sheet = len(ws)

	print('filename - ' + filename + ' / size - ' + str(size), file=fn)

	# print file infomation
	for i in range(number_sheet):
		sheet_name = wb.get_sheet_by_name(ws[i])
		sn = ws[i]
		print('	' + sn + ' - ' + \
			'rows : ' + str(sheet_name.max_row) + \
			' cols : ' + str(sheet_name.max_column), file=fn)

# make file name
folder = os.getcwd()
notice = "작업 위치 : "
print(notice + folder)
now = time.localtime()
tstamp = "_%02d-%02d-%02d_%02d-%02d-%02d" % (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
name = 'FILES_INFO'+tstamp+'.txt'

# count xls, xlsx files
filenames = os.listdir(folder)
xlfile_count = 0
for filename in filenames:
	if len(filename.split('.')) == 2:
		if filename.split('.')[1] == 'xls':
			xlfile_count += 1
		elif filename.split('.')[1] == 'xlsx':
			xlfile_count += 1

# no excel file process
if xlfile_count == 0:
	print('\n'+'작업할 폴더에 파일이 없습니다.')
	print('프로그램 사용 중 문제 발생 시 전산팀 최민우에게 연락 주세요.'+'\n')
	a = input('press enter plz')
	sys.exit(0)

fn = open(name, 'w')

# make report
filenames = os.listdir(folder)
for filename in filenames:
	if len(filename.split('.')) == 2:
		if filename.split('.')[1] == 'xls':
			print(filename)
			size = os.path.getsize(filename)
			search_xls(filename, fn, size)
			print("	", file=fn)
		elif filename.split('.')[1] == 'xlsx':
			print(filename)
			size = os.path.getsize(filename)
			search_xlsx(filename, fn, size)
			print("	", file=fn)

# make file tag
dline = '================================'
manage_massage = "파일 관리자"
m_manager = " 정 : "
s_manager = " 부 : "
fn.write(dline)
print('\n' + manage_massage + '\n' + m_manager + '\n' + s_manager, file=fn)
fn.write(dline+'\n'+'\n')

# file lod
ap = "변경 사항 기록"
fn.write(ap)
fn.close()

print ('\n'+'작업이 완료되었습니다.')
print ('프로그램 사용 중 문제 발생 시 전산팀 최민우에게 연락 주세요.'+'\n')

a = input('press enter plz')