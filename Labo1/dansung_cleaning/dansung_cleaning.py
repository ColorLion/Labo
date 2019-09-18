import xlrd
import os.path
import openpyxl

# 통계용 전역 변수
# [년도, 숫자]
STATIC_JA = [0, 0]
# [년도, allho, 1st, oth]
STATIC_HO = [0, 0, 0, 0]
# [년도, all, 1st, oth]
STATIC_FILE = [0, 0, 0, 0]
# 주호 저장용 변수
MAIN_HO = []

def change_fig(data):
    # 이순, 통, 호의 자리수 변경
    if type(data[8]) == float:  # 이순
        if len(str(data[8]).split('.')[0]) == 1:
            data[8] = "00" + str(data[8]).split('.')[0]
        elif len(str(data[8]).split('.')[0]) == 2:
            data[8] = "0" + str(data[8]).split('.')[0]
    else:
        data[8] = "10" + str(data[8]).replace("\'", "")

    if type(data[11]) == float:  # 통
        if len(str(data[11]).split('.')[0]) == 1:
            data[11] = "00" + str(data[11]).split('.')[0]
        elif len(str(data[11]).split('.')[0]) == 2:
            data[11] = "0" + str(data[11]).split('.')[0]
    else:
        data[11] = "10" + str(data[11]).replace("\'", "")

    if type(data[13]) == float:  # 호
        if len(str(data[13]).split('.')[0]) == 1:
            data[13] = "00" + str(data[13]).split('.')[0]
        elif len(str(data[8]).split('.')[0]) == 2:
            data[13] = "0" + str(data[13]).split('.')[0]
    else:
        data[13] = "10" + str(data[13]).replace("\'", "")

    # 변경된 자리수의 데이터 리턴
    return data

def extract_data(file, st_dir, oth_dir, sheet_static, sheet_ho, sheet_ja):
    # static 변수 초기화
    global STATIC_JA
    global STATIC_HO
    global STATIC_FILE
    global MAIN_HO

    # 연도 채우기
    STATIC_JA[0] = file.split('-')[0]
    STATIC_HO[0] = file.split('-')[0]
    STATIC_FILE[0] = file.split('-')[0]

    # 저장파일 객체 생성
    xlsx_1st = openpyxl.Workbook()
    xlsx_oth = openpyxl.Workbook()

    # 저장파일 오픈
    output_1st = xlsx_1st.active
    output_oth = xlsx_oth.active

    # 파일을 추출할 xls 파일 오픈
    target = xlrd.open_workbook(file, formatting_info=True)

    ws = target.sheet_by_index(0)

    st_row = ws.row_values(0)
    st_row.insert(0, "ID")
    output_1st.append(st_row)
    output_oth.append(st_row)
    for i in range(1, ws.nrows):
        # 원본 파일의 데이터 1줄을 읽어옮
        data = ws.row_values(i)

        # 리, 통, 호의 자릿수 변경
        data = change_fig(data)
        # 통계
        STATIC_FILE[1] += 1

        # 주호 저장
        if data[17] == "주호":
            STATIC_HO[1] += 1
            MAIN_HO = data

            if data[20] and data[22]:
                STATIC_HO[2] += 1
            else:
                STATIC_HO[3] += 1

        # 호내 위상이 '자'인 데이터의 '성(한자)', '성(한글)'부분을 채워 넣을 것
        if data[17] == "자" and MAIN_HO[20] and MAIN_HO[22]:
            STATIC_JA[1] += 1
            data[20] = MAIN_HO[20]
            data[22] = MAIN_HO[22]

        # 데이터 저장
        if MAIN_HO[20] and MAIN_HO[22]:
            STATIC_FILE[2] += 1
            data.insert(0, STATIC_HO[1])
            output_1st.append(data)
            del data[0]
        else:
            STATIC_FILE[3] += 1
            data.insert(0, STATIC_HO[1])
            output_oth.append(data)
            del data[0]

    # 1st data save
    save_file_1st = st_dir + file.split('.')[0] + "_1st.xlsx"
    xlsx_1st.save(save_file_1st)

    # other data save
    save_file_oth = oth_dir + file.split('.')[0] + "_oth.xlsx"
    xlsx_oth.save(save_file_oth)

    # static data save
    sheet_ho.append(STATIC_HO)
    sheet_ja.append(STATIC_JA)
    sheet_static.append(STATIC_FILE)

    # static 전역변수 초기화
    STATIC_HO = [0, 0, 0, 0]
    STATIC_JA = [0, 0]
    STATIC_FILE = [0, 0, 0, 0]

def main():
    # 작업 위치 및 저장 위치 정의
    work_dir = os.getcwd()
    print(work_dir + "현재 작업 위치")
    root_dir = work_dir + '\\' + "output" + '\\'
    st_dir = root_dir + '1st' + '\\'
    oth_dir = root_dir + 'oth' + '\\'

    # output 저장 폴더 확인
    if os.path.isdir(root_dir) == 0:
        os.mkdir(root_dir)
        os.mkdir(st_dir)
        os.mkdir(oth_dir)

    files = os.listdir(work_dir)

    # 통계 저장 파일 객체 생성
    xlsx_static = openpyxl.Workbook("JA")

    sheet_static = xlsx_static.create_sheet()
    sheet_static.title = "STATIC"
    sheet_static.append(["연도", "full", "1st", "oth"])

    sheet_ho = xlsx_static.create_sheet()
    sheet_ho.title = "HO"
    sheet_ho.append(["연도", "전체 호", "성o-1st", "성x-oth"])

    sheet_ja = xlsx_static.create_sheet()
    sheet_ja.title = "JA"
    sheet_ja.append(["연도", "성이 붙은 자"])

    # 폴더 내에 있는 xls 파일을 순차적으로 읽음
    for file in files:
        if len(file.split('.')) == 2:
            if file.split('.')[1] == 'xls':
                print(file)
                extract_data(file, st_dir, oth_dir, sheet_static, sheet_ho, sheet_ja)

    # static data save
    save_file_static = root_dir + '\\' + "cleaning_static.xlsx"
    xlsx_static.save(save_file_static)

main()