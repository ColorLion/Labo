import os.path
import openpyxl

# for Statics
main_ho = []
id = []
hoid = 0

def extract_data(i, save_file):
    global main_ho
    global hoid
    # 추출 데이터
        # 성/명(한자, 한글), 출생년도, 간지, 년도, 이순, 통, 호 면명, 리명
    a = [hoid]
    a.append(i[4].value)            # 년도
    a.append(i[6].value)            # 면명
    a.append(i[10].value)           # 리명
    a.append(i[8].value)            # 이순
    a.append(i[11].value)           # 통
    a.append(i[13].value)           # 호
    a.append(i[14].value)           # 주호(한자)
    a.append(i[15].value)           # 주호(한글)
    a.append(i[20].value)           # 성(한자)
    a.append(i[21].value)           # 명(한자)
    a.append(i[22].value)           # 성(한글)
    a.append(i[23].value)           # 명(한글)
    a.append(i[17].value)           # 호내위상
    if type(i[24].value) == int:    # 출생년도
        a.append(i[4].value - i[24].value)
    else:
        a.append(i[24].value)
    a.append(i[25].value)           # 간지(한자)
    a.append(i[26].value)           # 간지(한글)

    if i[17].value == "주호":
        hoid += 1
        a[0] = hoid
        main_ho = a

    if i[17].value == "자":
        a[9] = main_ho[9]
        a[11] = main_ho[11]
    save_file.append(a)

def gather_static(i, job_ho_info, main_ho_info):
    # 기본 통계
    job_ho_info[1] += 1
    job = i[18].value
    if job == None:
        job_ho_info[4] += 1
    elif '奴' in job or '婢' in job:
        job_ho_info[3] += 1
    elif 'x' in job:
        job_ho_info[5] += 1
    else:
        job_ho_info[2] += 1
        
    # 호 관련 통계
    if i[17].value == "주호":
        main_ho_info[1] += 1
        if job == None:
            main_ho_info[4] += 1
        elif '奴' in job or '婢' in job:
            main_ho_info[3] += 1
        elif 'x' in job:
            main_ho_info[5] += 1
        else:
            main_ho_info[2] += 1

def extract_xlsx(file, main_hoid_output, static_output1, static_output2):
    # target file open
    xlsx = openpyxl.load_workbook(filename=file, data_only=True)
    sheet = xlsx['Sheet']

    # static_first1 = ["file_name", "전체 인물 수", "노비가 아닌 사람 수", "노비", "직역이 빈칸인 사람 수", "직역에 x가 포함된 사람 수"]
    job_ho_info = [file.split('-')[0], 0, 0, 0, 0, 0]
    # static_first2 = ["연도", "전체 호의 수", "노비가 아닌 주호의 수", "노비인 주호의 수", "직역이 비어있는 주호", "직역에 x가 포함된 주호"]
    main_ho_info = [file.split('-')[0], 0, 0, 0, 0, 0]

    for i in sheet.rows:
        # 첫 번째 줄 처리
        if i[0].value == "原本":
            continue
        else:
            extract_data(i, main_hoid_output)
        # Static
        gather_static(i, job_ho_info, main_ho_info)

    static_output1.append(job_ho_info)
    static_output2.append(main_ho_info)

def main():
    # 작업 위치 및 저장 위치 정의
    work_dir = os.getcwd()
    print(work_dir + "현재 작업 위치")
    main_id_dir = work_dir + '\\' + "output_ho" + '\\'
    static_id_dir = work_dir + '\\' + "output_static" + '\\'

    save_file_ho = main_id_dir + "ho_id.xlsx"
    save_file_static = static_id_dir + "statics.xlsx"

    # report나올 것
    xlsx1 = openpyxl.Workbook()
    ho_first = ["id", "연도", "면명", "리명", "이순", "통", "호", "주호(한자)", "주호(한글)", "성(한자)", \
               "명(한자)", "성(한글)", "명(한글)", "호내위상", "출생년도", "간지(한자)", "간지(한글)"]
    main_hoid_output = xlsx1.active
    main_hoid_output.append(ho_first)

    # 통계 자료 출력용
    xlsx2 = openpyxl.Workbook()
    static_output1 = xlsx2.active
    static_output1.title = "기본 통계자료"
    static_first1 = ["연도", "전체 인물 수", "노비가 아닌 사람 수", "노비", "직역이 빈칸인 사람 수", "직역에 x가 포함된 사람 수"]
    static_output2 = xlsx2.create_sheet()
    static_output2.title = "호 통계 자료"
    static_first2 = ["연도", "전체 호의 수", "노비가 아닌 주호의 수", "노비인 주호의 수", "직역이 비어있는 주호", "직역에 x가 포함된 주호"]
    static_output1.append(static_first1)
    static_output2.append(static_first2)

    # output 저장 폴더 확인
    if os.path.isdir(main_id_dir) == 0:
        os.mkdir(main_id_dir)
    if os.path.isdir(static_id_dir) == 0:
        os.mkdir(static_id_dir)

    files = os.listdir(work_dir)

    for file in files:
        if len(file.split('.')) == 2:
            if file.split('.')[1] == 'xlsx':
                print(file)
                extract_xlsx(file, main_hoid_output, static_output1, static_output2)

    xlsx1.save(save_file_ho)
    xlsx2.save(save_file_static)

main()